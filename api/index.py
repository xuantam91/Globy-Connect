from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import select, func
import zipfile
import xml.etree.ElementTree as ET
import io

def parse_xlsx_built_in(file_content):
    try:
        with zipfile.ZipFile(io.BytesIO(file_content)) as z:
            shared_strings = []
            if "xl/sharedStrings.xml" in z.namelist():
                ss_xml = z.read("xl/sharedStrings.xml")
                root = ET.fromstring(ss_xml)
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for t in root.findall('.//ns:t', ns):
                    shared_strings.append(t.text or "")
            
            sheet_xml = z.read("xl/worksheets/sheet1.xml")
            root = ET.fromstring(sheet_xml)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            rows = []
            for row in root.findall('.//ns:row', ns):
                row_cells = []
                for c in row.findall('ns:c', ns):
                    v_elem = c.find('ns:v', ns)
                    v = v_elem.text if v_elem is not None else ""
                    t = c.get('t')
                    if t == 's' and v:
                        idx = int(v)
                        v = shared_strings[idx] if idx < len(shared_strings) else ""
                    row_cells.append(v)
                rows.append(row_cells)
            return rows
    except Exception as e:
        print(f"Error parsing xlsx natively: {e}")
        return []

from typing import List, Optional
import json
from pydantic import BaseModel

from api.app.core.database import Base, engine, get_db
from api.app.models import Device, XiaozhiAccount, Preset, SystemSetting
from api.app.services.xiaozhi_service import XiaozhiService

# Auto create tables on startup
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Xiaozhi Lite Connect API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def seed_database():
    db = next(get_db())
    try:
        preset_count = db.scalar(select(func.count(Preset.id)))
        if preset_count == 0:
            default_presets = [
                Preset(
                    name="Học tiếng Anh với Lily",
                    description="Lily - Trợ lý giúp bé học và luyện giao tiếp tiếng Anh tự nhiên.",
                    llm_model="qwen",
                    language="en",
                    tts_voice="zh_female_mengyatou_mars_bigtts",
                    tts_speech_speed="normal",
                    asr_speed="normal",
                    tts_pitch=0,
                    mcp_endpoints_json='["2", "104"]',
                    character_prompt="You are Lily, a kind and patient English teacher for children. Help children learn English in a fun way.",
                    is_public=True
                ),
                Preset(
                    name="Kể chuyện cùng Thỏ Ngọc",
                    description="Thỏ Ngọc - Giọng nói ấm áp chuyên kể truyện cổ tích và bài học đạo đức.",
                    llm_model="deepseek",
                    language="vi",
                    tts_voice="vi-VN-HoaiMyNeural",
                    tts_speech_speed="normal",
                    asr_speed="normal",
                    tts_pitch=0,
                    mcp_endpoints_json='["2", "104"]',
                    character_prompt="Bạn là Thỏ Ngọc, người bạn kể chuyện cổ tích cực kỳ truyền cảm, ấm áp cho trẻ em Việt Nam.",
                    is_public=True
                ),
                Preset(
                    name="Gia sư Toán Cú Thông Thái",
                    description="Cú Thông Thái - Giúp trẻ giải các bài toán tư duy đố vui sinh động.",
                    llm_model="qwen",
                    language="vi",
                    tts_voice="vi-VN-NamMinhNeural",
                    tts_speech_speed="normal",
                    asr_speed="normal",
                    tts_pitch=0,
                    mcp_endpoints_json='["2", "104"]',
                    character_prompt="Bạn là gia sư toán học Cú Thông Thái. Hãy kiên nhẫn giảng giải các phép tính đố vui tiểu học cho trẻ.",
                    is_public=True
                )
            ]
            db.add_all(default_presets)
            db.commit()
            print("Successfully seeded default presets!")
    except Exception as e:
        print(f"Error seeding database: {e}")
    finally:
        db.close()

# Pydantic schemas
class AccountCreate(BaseModel):
    label: str
    bearer_token: str

class BulkConfigApply(BaseModel):
    device_ids: List[int]
    llm_model: str
    language: str
    tts_voice: str
    tts_speech_speed: str
    asr_speed: str
    tts_pitch: int
    mcp_endpoints: List[str]
    character: str

def detect_gender(name: str, tts_voice: str = "") -> str:
    """Auto-detect gender from preset name or voice ID."""
    n = (name or "").lower()
    v = (tts_voice or "").lower()
    if any(k in n for k in ["(nữ)", "(female)", "nữ", "female"]):
        return "female"
    if any(k in n for k in ["(nam)", "(male)", "nam", "male"]):
        return "male"
    # Fallback: check voice_id
    if "female" in v or "girl" in v or "lady" in v:
        return "female"
    if "male" in v and "female" not in v:
        return "male"
    if "man" in v or "boy" in v:
        return "male"
    return "neutral"

class PresetCreate(BaseModel):
    name: str
    description: Optional[str] = None
    llm_model: str
    language: str
    tts_voice: str
    tts_speech_speed: str = "normal"
    asr_speed: str = "normal"
    tts_pitch: int = 0
    mcp_endpoints: List[str] = ["2", "104"]
    character_prompt: str
    icon: Optional[str] = "✨"
    version: Optional[int] = 1
    gender: Optional[str] = None  # 'male', 'female', 'neutral' - auto-detected if not provided

class QrApply(BaseModel):
    mac_address: str
    template_id: str
    language: Optional[str] = None
    llm_model: Optional[str] = None
    tts_voice: Optional[str] = None
    tts_speech_speed: Optional[str] = None
    asr_speed: Optional[str] = None
    tts_pitch: Optional[int] = None
    mcp_endpoints: Optional[List[str]] = None
    character: Optional[str] = None
    contribute: Optional[bool] = False

class DeviceActivate(BaseModel):
    code: str
    name: Optional[str] = None

class AdminLogin(BaseModel):
    password: str

class AdminChangePassword(BaseModel):
    current_password: str
    new_password: str


@app.get("/api/health")
def health():
    return {"status": "healthy", "service": "xiaozhi-lite-connect"}

# Accounts API
@app.get("/api/accounts")
def get_accounts(db: Session = Depends(get_db)):
    accounts = db.scalars(select(XiaozhiAccount)).all()
    return {
        "success": True, 
        "data": [{"id": a.id, "label": a.label, "bearer_token": a.bearer_token, "is_active": a.is_active} for a in accounts]
    }

@app.post("/api/accounts")
def create_account(acc: AccountCreate, db: Session = Depends(get_db)):
    new_acc = XiaozhiAccount(label=acc.label, bearer_token=acc.bearer_token)
    db.add(new_acc)
    db.commit()
    db.refresh(new_acc)
    return {"success": True, "data": {"id": new_acc.id, "label": new_acc.label}}

@app.delete("/api/accounts/{account_id}")
def delete_account(account_id: int, db: Session = Depends(get_db)):
    acc = db.scalar(select(XiaozhiAccount).where(XiaozhiAccount.id == account_id))
    if not acc:
        raise HTTPException(status_code=404, detail="Account not found")
    db.delete(acc)
    db.commit()
    return {"success": True}

@app.delete("/api/devices/{device_id}")
def delete_device(device_id: int, db: Session = Depends(get_db)):
    device = db.scalar(select(Device).where(Device.id == device_id))
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    # 1. Delete from Supabase first if configured
    try:
        from api.app.core.database import SUPABASE_URL, SUPABASE_KEY
        import httpx
        if SUPABASE_URL and SUPABASE_KEY:
            supabase_headers = {
                "apikey": SUPABASE_KEY,
                "Authorization": f"Bearer {SUPABASE_KEY}",
                "Content-Type": "application/json"
            }
            res = httpx.delete(
                f"{SUPABASE_URL}/rest/v1/devices?external_id=eq.{device.external_id}",
                headers=supabase_headers
            )
            if res.status_code not in (200, 204, 404):
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to delete device from Supabase (Status: {res.status_code}): {res.text}")
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting device from Supabase: {e}")

    # 2. Delete from local SQLite
    db.delete(device)
    db.commit()
    return {"success": True, "message": "Xóa thiết bị khỏi Globy thành công!"}

# Devices API
@app.get("/api/devices")
def get_devices(db: Session = Depends(get_db)):
    devices = db.scalars(select(Device).order_by(Device.name)).all()
    # Check if empty, auto-create mock devices for first-time use
    if not devices:
        # Simple check, if no accounts exists, create a default mock account to run demo
        acc = db.scalar(select(XiaozhiAccount))
        if not acc:
            acc = XiaozhiAccount(label="Mock Demo Token", bearer_token="mock")
            db.add(acc)
            db.commit()
            db.refresh(acc)
        
        # Trigger mock sync
        service = XiaozhiService(db)
        import asyncio
        asyncio.run(service.sync_all_accounts())
        devices = db.scalars(select(Device).order_by(Device.name)).all()

    return {
        "success": True,
        "data": [
            {
                "id": d.id,
                "external_id": d.external_id,
                "name": d.name,
                "status": d.status,
                "mac_address": d.mac_address,
                "current_language": d.current_language,
                "current_voice": d.current_voice,
                "ai_prompt_template": d.ai_prompt_template,
                "asr_speed": d.asr_speed,
                "tts_speech_speed": d.tts_speech_speed,
                "tts_pitch": d.tts_pitch,
                "mcp_endpoints": json.loads(d.mcp_endpoints_json) if d.mcp_endpoints_json else ["2", "104"],
                "last_seen_at": d.last_seen_at.isoformat() if d.last_seen_at else None,
                "board_name": d.board_name,
                "serial_number": d.serial_number,
                "is_auth": d.is_auth,
                "app_version": d.app_version,
                "account_label": d.account.label if d.account else "Chưa liên kết",
                "llm_model": d.llm_model or "qwen"
            }
            for d in devices
        ]
    }


@app.get("/api/devices/by-mac")
def get_device_by_mac(mac: str = Query(...), db: Session = Depends(get_db)):
    cleaned_mac = mac.replace(":", "").replace("-", "").strip().lower()
    formatted_mac = ":".join(cleaned_mac[i:i+2] for i in range(0, 12, 2)) if len(cleaned_mac) == 12 else cleaned_mac
    
    device = db.scalar(
        select(Device).where(
            (Device.mac_address.ilike(formatted_mac)) | 
            (Device.mac_address.ilike(mac.strip())) |
            (Device.external_id == mac.strip())
        )
    )
    if not device:
        return {
            "success": False,
            "message": "Không tìm thấy loa có địa chỉ MAC này trên hệ thống. Vui lòng kiểm tra lại thiết bị hoặc tiến hành đồng bộ tài khoản Xiaozhi."
        }
        
    return {
        "success": True,
        "data": {
            "id": device.id,
            "external_id": device.external_id,
            "name": device.name,
            "status": device.status,
            "mac_address": device.mac_address,
            "current_language": device.current_language,
            "current_voice": device.current_voice,
            "ai_prompt_template": device.ai_prompt_template,
            "asr_speed": device.asr_speed,
            "tts_speech_speed": device.tts_speech_speed,
            "tts_pitch": device.tts_pitch,
            "mcp_endpoints": json.loads(device.mcp_endpoints_json) if device.mcp_endpoints_json else ["2", "104"]
        }
    }

@app.post("/api/devices/sync")
@app.get("/api/devices/sync")
async def sync_devices(db: Session = Depends(get_db)):
    service = XiaozhiService(db)
    count = await service.sync_all_accounts()
    return {"success": True, "records_fetched": count}

@app.post("/api/devices/config")
async def apply_bulk_config(cfg: BulkConfigApply, db: Session = Depends(get_db)):
    service = XiaozhiService(db)
    devices = db.scalars(select(Device).where(Device.id.in_(cfg.device_ids))).all()
    
    if not devices:
        raise HTTPException(status_code=404, detail="No devices found")
        
    success_count = 0
    for device in devices:
        ok = await service.update_device_config(
            device=device,
            llm_model=cfg.llm_model,
            language=cfg.language,
            tts_voice=cfg.tts_voice,
            tts_speech_speed=cfg.tts_speech_speed,
            asr_speed=cfg.asr_speed,
            tts_pitch=cfg.tts_pitch,
            mcp_endpoints=cfg.mcp_endpoints,
            character=cfg.character
        )
        if ok:
            success_count += 1
            
    return {"success": True, "updated_count": success_count}

# QR scanning Quick config API
@app.post("/api/qr/apply")
async def apply_qr_template(qr: QrApply, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    cleaned_mac = qr.mac_address.replace(":", "").replace("-", "").strip().lower()
    formatted_mac = ":".join(cleaned_mac[i:i+2] for i in range(0, 12, 2)) if len(cleaned_mac) == 12 else cleaned_mac

    device = db.scalar(
        select(Device).where(
            (Device.mac_address.ilike(formatted_mac)) | 
            (Device.mac_address.ilike(qr.mac_address.strip())) |
            (Device.external_id == qr.mac_address.strip())
        )
    )
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
        
    if qr.template_id == "custom":
        if not qr.language or not qr.llm_model or not qr.tts_voice or not qr.character:
            raise HTTPException(status_code=400, detail="Missing required custom configuration fields")
            
        mcp_endpoints = qr.mcp_endpoints if qr.mcp_endpoints is not None else ["2", "104"]
        asr_speed = qr.asr_speed or "normal"
        tts_speech_speed = qr.tts_speech_speed or "normal"
        tts_pitch = qr.tts_pitch if qr.tts_pitch is not None else 0
        
        service = XiaozhiService(db)
        ok = await service.update_device_config(
            device=device,
            llm_model=qr.llm_model,
            language=qr.language,
            tts_voice=qr.tts_voice,
            tts_speech_speed=tts_speech_speed,
            asr_speed=asr_speed,
            tts_pitch=tts_pitch,
            mcp_endpoints=mcp_endpoints,
            character=qr.character
        )

        if ok and qr.contribute:
            preset_name = f"User Config ({qr.mac_address[-4:] if len(qr.mac_address) >= 4 else qr.mac_address})"
            existing = db.scalar(select(Preset).where(Preset.name == preset_name))
            if not existing:
                db_preset = Preset(
                    name=preset_name,
                    description=f"Cấu hình đóng góp bởi người dùng thiết bị {qr.mac_address}",
                    llm_model=qr.llm_model,
                    language=qr.language,
                    tts_voice=qr.tts_voice,
                    tts_speech_speed=tts_speech_speed,
                    asr_speed=asr_speed,
                    tts_pitch=tts_pitch,
                    mcp_endpoints_json=json.dumps(mcp_endpoints),
                    character_prompt=qr.character,
                    is_public=True
                )
                db.add(db_preset)
                db.commit()
    else:
        # 1. First search in Database Presets
        selected_tmpl = None
        try:
            preset_id_int = int(qr.template_id)
            db_preset = db.scalar(select(Preset).where(Preset.id == preset_id_int))
            if db_preset:
                selected_tmpl = {
                    "llm_model": db_preset.llm_model,
                    "language": db_preset.language,
                    "tts_voice": db_preset.tts_voice,
                    "tts_speech_speed": db_preset.tts_speech_speed,
                    "asr_speed": db_preset.asr_speed,
                    "tts_pitch": db_preset.tts_pitch,
                    "mcp_endpoints": json.loads(db_preset.mcp_endpoints_json) if db_preset.mcp_endpoints_json else ["2", "104"],
                    "character": db_preset.character_prompt
                }
        except ValueError:
            pass # Not an integer ID, try hardcoded template

        if not selected_tmpl:
            # 2. Fallback to hardcoded templates
            templates = {
                "english_lily": {
                    "llm_model": "qwen",
                    "language": "en",
                    "tts_voice": "zh_female_shuangkuaisisi_moon_bigtts", # Skye
                    "tts_speech_speed": "normal",
                    "asr_speed": "slow",
                    "tts_pitch": 0,
                    "mcp_endpoints": ["2", "104"],
                    "character": "Your name is Lily. You are a friendly, patient, and native English teacher.\nSpeak in short, clear, and simple sentences.\nAlways keep the conversation going with a follow-up question."
                },
                "vietnamese_story": {
                    "llm_model": "xz-lite",
                    "language": "vi",
                    "tts_voice": "vi-VN-HoaiMyNeural", # Hoai My
                    "tts_speech_speed": "normal",
                    "asr_speed": "normal",
                    "tts_pitch": 0,
                    "mcp_endpoints": ["2", "104"],
                    "character": "Bạn là cô tiên kể chuyện cổ tích. Giọng nói của bạn ấm áp và lôi cuốn.\nHãy kể những câu chuyện cổ tích Việt Nam ngắn gọn cho trẻ em nghe."
                },
                "bilingual_tutor": {
                    "llm_model": "deepseek",
                    "language": "en",
                    "tts_voice": "zh_male_wennuanahu_moon_bigtts", # Alvin
                    "tts_speech_speed": "normal",
                    "asr_speed": "normal",
                    "tts_pitch": 0,
                    "mcp_endpoints": ["2", "104"],
                    "character": "You are a bilingual English-Vietnamese AI tutor. You translate phrases and answer science/history questions simply for kids in both languages."
                }
            }
            selected_tmpl = templates.get(qr.template_id)

        if not selected_tmpl:
            raise HTTPException(status_code=400, detail="Invalid template ID")
            
        service = XiaozhiService(db)
        ok = await service.update_device_config(
            device=device,
            llm_model=selected_tmpl["llm_model"],
            language=selected_tmpl["language"],
            tts_voice=selected_tmpl["tts_voice"],
            tts_speech_speed=selected_tmpl["tts_speech_speed"],
            asr_speed=selected_tmpl["asr_speed"],
            tts_pitch=selected_tmpl["tts_pitch"],
            mcp_endpoints=selected_tmpl["mcp_endpoints"],
            character=selected_tmpl["character"]
        )
    
    if ok:
        queue_supabase_sync(device, db, background_tasks)
        return {"success": True}
    else:
        return {"success": False, "message": "Failed to push configuration to Xiaozhi API"}

# Presets APIs
@app.get("/api/presets")
def get_presets(db: Session = Depends(get_db)):
    presets_db = db.scalars(select(Preset).order_by(Preset.sort_order.asc(), Preset.id.asc())).all()
    res = []
    for p in presets_db:
        res.append({
            "id": str(p.id),
            "name": p.name,
            "description": p.description,
            "llm_model": p.llm_model,
            "language": p.language,
            "tts_voice": p.tts_voice,
            "tts_speech_speed": p.tts_speech_speed,
            "asr_speed": p.asr_speed,
            "tts_pitch": p.tts_pitch,
            "mcp_endpoints": json.loads(p.mcp_endpoints_json) if p.mcp_endpoints_json else ["2", "104"],
            "character": p.character_prompt,
            "icon": p.icon or "✨",
            "version": p.version or 1,
            "is_public": p.is_public,
            "sort_order": p.sort_order or 0,
            "is_default": p.is_default or False,
            "gender": p.gender or "neutral"
        })
    return {"success": True, "data": res}

@app.post("/api/presets")
def create_preset(p: PresetCreate, db: Session = Depends(get_db)):
    resolved_gender = p.gender if p.gender else detect_gender(p.name, p.tts_voice)
    db_preset = Preset(
        name=p.name,
        description=p.description,
        llm_model=p.llm_model,
        language=p.language,
        tts_voice=p.tts_voice,
        tts_speech_speed=p.tts_speech_speed,
        asr_speed=p.asr_speed,
        tts_pitch=p.tts_pitch,
        mcp_endpoints_json=json.dumps(p.mcp_endpoints),
        character_prompt=p.character_prompt,
        icon=p.icon or "✨",
        version=p.version or 1,
        gender=resolved_gender,
        is_public=True
    )
    db.add(db_preset)
    db.commit()
    db.refresh(db_preset)
    return {"success": True, "id": db_preset.id}

@app.put("/api/presets/{preset_id}")
def update_preset(preset_id: int, p: PresetCreate, db: Session = Depends(get_db)):
    db_preset = db.scalar(select(Preset).where(Preset.id == preset_id))
    if not db_preset:
        return {"success": False, "message": "Preset not found"}
    
    db_preset.name = p.name
    db_preset.description = p.description
    db_preset.llm_model = p.llm_model
    db_preset.language = p.language
    db_preset.tts_voice = p.tts_voice
    db_preset.tts_speech_speed = p.tts_speech_speed
    db_preset.asr_speed = p.asr_speed
    db_preset.tts_pitch = p.tts_pitch
    db_preset.mcp_endpoints_json = json.dumps(p.mcp_endpoints)
    db_preset.character_prompt = p.character_prompt
    if p.icon:
        db_preset.icon = p.icon
    
    # Update gender - use explicit value or auto-detect
    db_preset.gender = p.gender if p.gender else detect_gender(p.name, p.tts_voice)
    
    # Auto-increment version when edited
    db_preset.version = (db_preset.version or 1) + 1
    
    db.commit()
    db.refresh(db_preset)
    return {
        "success": True, 
        "data": {
            "id": str(db_preset.id),
            "name": db_preset.name,
            "description": db_preset.description,
            "llm_model": db_preset.llm_model,
            "language": db_preset.language,
            "tts_voice": db_preset.tts_voice,
            "tts_speech_speed": db_preset.tts_speech_speed,
            "asr_speed": db_preset.asr_speed,
            "tts_pitch": db_preset.tts_pitch,
            "mcp_endpoints": json.loads(db_preset.mcp_endpoints_json) if db_preset.mcp_endpoints_json else ["2", "104"],
            "character": db_preset.character_prompt,
            "icon": db_preset.icon,
            "version": db_preset.version,
            "is_public": db_preset.is_public,
            "gender": db_preset.gender or "neutral"
        }
    }

@app.delete("/api/presets/{preset_id}")
def delete_preset(preset_id: int, db: Session = Depends(get_db)):
    db_preset = db.scalar(select(Preset).where(Preset.id == preset_id))
    if not db_preset:
        return {"success": False, "message": "Preset not found"}
    db.delete(db_preset)
    db.commit()
    return {"success": True}

@app.post("/api/presets/reorder")
def reorder_presets(body: dict, db: Session = Depends(get_db)):
    """Receive ordered list of preset IDs and update sort_order accordingly."""
    ordered_ids = body.get("ordered_ids", [])
    if not ordered_ids:
        return {"success": False, "message": "No ordered_ids provided"}
    for idx, pid in enumerate(ordered_ids):
        db_preset = db.scalar(select(Preset).where(Preset.id == int(pid)))
        if db_preset:
            db_preset.sort_order = idx
    db.commit()
    return {"success": True}

@app.post("/api/presets/{preset_id}/set-default")
def set_default_preset(preset_id: int, db: Session = Depends(get_db)):
    """Set a preset as the default. Clears default flag from all others."""
    # Clear all defaults first
    all_presets = db.scalars(select(Preset)).all()
    for p in all_presets:
        p.is_default = False
    # Set the target as default
    target = db.scalar(select(Preset).where(Preset.id == preset_id))
    if not target:
        return {"success": False, "message": "Preset not found"}
    target.is_default = True
    db.commit()
    return {"success": True, "default_preset_id": str(preset_id)}

@app.post("/api/presets/import")
async def import_presets(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = file.filename.lower()
    content = await file.read()
    
    imported_count = 0
    if filename.endswith(".json"):
        try:
            data = json.loads(content.decode("utf-8"))
            if isinstance(data, dict):
                data = [data]
            for item in data:
                db_preset = Preset(
                    name=item.get("name", "Unnamed Preset"),
                    description=item.get("description", ""),
                    llm_model=item.get("llm_model", "qwen"),
                    language=item.get("language", "en"),
                    tts_voice=item.get("tts_voice", ""),
                    tts_speech_speed=item.get("tts_speech_speed", "normal"),
                    asr_speed=item.get("asr_speed", "normal"),
                    tts_pitch=int(item.get("tts_pitch", 0)),
                    mcp_endpoints_json=json.dumps(item.get("mcp_endpoints", ["2", "104"])),
                    character_prompt=item.get("character", item.get("character_prompt", "")),
                    icon=item.get("icon", "✨"),
                    version=int(item.get("version", 1)),
                    gender=item.get("gender") or detect_gender(item.get("name", ""), item.get("tts_voice", "")),
                    is_public=True
                )
                db.add(db_preset)
                imported_count += 1
            db.commit()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Lỗi phân tích JSON: {str(e)}")
            
    elif filename.endswith(".csv"):
        try:
            import csv
            import io
            text = content.decode("utf-8")
            f = io.StringIO(text)
            reader = csv.DictReader(f)
            for row in reader:
                endpoints_str = row.get("mcp_endpoints", "2,104")
                if "[" in endpoints_str:
                    try:
                        mcp_endpoints = json.loads(endpoints_str.replace("'", '"'))
                    except:
                        mcp_endpoints = ["2", "104"]
                else:
                    mcp_endpoints = [x.strip() for x in endpoints_str.split(",") if x.strip()]
                
                try:
                    ver_val = int(row.get("version", 1))
                except:
                    ver_val = 1

                db_preset = Preset(
                    name=row.get("name", "Unnamed Preset"),
                    description=row.get("description", ""),
                    llm_model=row.get("llm_model", "qwen"),
                    language=row.get("language", "en"),
                    tts_voice=row.get("tts_voice", ""),
                    tts_speech_speed=row.get("tts_speech_speed", "normal"),
                    asr_speed=row.get("asr_speed", "normal"),
                    tts_pitch=int(row.get("tts_pitch", 0)),
                    mcp_endpoints_json=json.dumps(mcp_endpoints),
                    character_prompt=row.get("character", row.get("character_prompt", "")),
                    icon=row.get("icon", "✨"),
                    version=ver_val,
                    gender=row.get("gender") or detect_gender(row.get("name", ""), row.get("tts_voice", "")),
                    is_public=True
                )
                db.add(db_preset)
                imported_count += 1
            db.commit()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Lỗi phân tích CSV: {str(e)}")
            
    elif filename.endswith(".xlsx"):
        try:
            import openpyxl
            import io as io
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = []
            for r in ws.iter_rows(values_only=True):
                rows.append([str(val) if val is not None else "" for val in r])
            
            if not rows or len(rows) < 2:
                raise Exception("File Excel rỗng hoặc thiếu tiêu đề cột.")
            
            headers = [h.strip().lower() if h else "" for h in rows[0]]
            col_map = {h: i for i, h in enumerate(headers) if h}
            
            for row in rows[1:]:
                if not row or len(row) == 0:
                    continue
                
                def get_cell(name, default=""):
                    idx = col_map.get(name)
                    return row[idx] if (idx is not None and idx < len(row)) else default
                
                name = get_cell("name", "")
                if not name:
                    continue
                    
                endpoints_str = get_cell("mcp_endpoints", "2,104")
                if "[" in endpoints_str:
                    try:
                        mcp_endpoints = json.loads(endpoints_str.replace("'", '"'))
                    except:
                        mcp_endpoints = ["2", "104"]
                else:
                    mcp_endpoints = [x.strip() for x in endpoints_str.split(",") if x.strip()]
                
                try:
                    pitch = int(float(get_cell("tts_pitch", 0)))
                except:
                    pitch = 0
                    
                try:
                    ver_val = int(float(get_cell("version", 1)))
                except:
                    ver_val = 1

                db_preset = Preset(
                    name=name,
                    description=get_cell("description", ""),
                    llm_model=get_cell("llm_model", "qwen"),
                    language=get_cell("language", "en"),
                    tts_voice=get_cell("tts_voice", ""),
                    tts_speech_speed=get_cell("tts_speech_speed", "normal"),
                    asr_speed=get_cell("asr_speed", "normal"),
                    tts_pitch=pitch,
                    mcp_endpoints_json=json.dumps(mcp_endpoints),
                    character_prompt=get_cell("character_prompt", get_cell("character", "")),
                    icon=get_cell("icon", "✨"),
                    version=ver_val,
                    gender=get_cell("gender", "") or detect_gender(name, get_cell("tts_voice", "")),
                    is_public=True
                )
                db.add(db_preset)
                imported_count += 1
            db.commit()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Lỗi phân tích Excel (.xlsx): {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Định dạng file không được hỗ trợ. Chỉ nhận .json, .csv, .xlsx")
        
    return {"success": True, "imported_count": imported_count}

@app.get("/api/presets/template")
def get_presets_template():
    import io
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Preset Template"
    
    # Headers
    headers = [
        "icon", "name", "description", "llm_model", "language", 
        "tts_voice", "tts_speech_speed", "asr_speed", "tts_pitch", 
        "mcp_endpoints", "character_prompt", "version"
    ]
    ws.append(headers)
    
    # Example Row
    example = [
        "📚", "Kể chuyện cổ tích", "Kể chuyện cổ tích Việt Nam",
        "qwen", "vi", "Vietnamese_kindhearted_girl", "normal", "normal", 0,
        "2,104", "Bạn là một người kể chuyện cổ tích thân thiện...", 1
    ]
    ws.append(example)
    
    # Add dropdown validations
    # 1. Models
    dv_model = DataValidation(type="list", formula1='"qwen,deepseek,doubao,xz-lite"', allow_blank=True)
    ws.add_data_validation(dv_model)
    dv_model.add("D2:D100")
    
    # 2. Languages
    dv_lang = DataValidation(type="list", formula1='"vi,en,zh,ja"', allow_blank=True)
    ws.add_data_validation(dv_lang)
    dv_lang.add("E2:E100")
    
    # 3. Voices
    all_voices = [
        "Vietnamese_kindhearted_girl", "Vietnamese_Serene_Man",
        "zh_female_shuangkuaisisi_moon_bigtts", "zh_male_jingqiangkanye_moon_bigtts",
        "zh_female_mengyatou_mars_bigtts", "zh_male_shaonianzixin_moon_bigtts",
        "zh_male_wennuanahu_moon_bigtts", "zh_female_wanwanxiaohe_moon_bigtts",
        "zh_female_linjianvhai_moon_bigtts", "zh_male_yuanboxiaoshu_moon_bigtts",
        "ja_female_1_v1", "ja_male_1_v1"
    ]
    voices_formula = f'"{",".join(all_voices)}"'
    dv_voice = DataValidation(type="list", formula1=voices_formula, allow_blank=True)
    ws.add_data_validation(dv_voice)
    dv_voice.add("F2:F100")
    
    # 4. Speeds (ASR and TTS)
    dv_speed = DataValidation(type="list", formula1='"slow,normal,fast"', allow_blank=True)
    ws.add_data_validation(dv_speed)
    dv_speed.add("G2:G100")
    dv_speed.add("H2:H100")
    
    # 5. Pitch
    dv_pitch = DataValidation(type="list", formula1='"-2,0,2"', allow_blank=True)
    ws.add_data_validation(dv_pitch)
    dv_pitch.add("I2:I100")
    
    # Save workbook to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=presets_template.xlsx"}
    )

class SettingsUpdate(BaseModel):
    supabase_url: Optional[str] = None
    supabase_key: Optional[str] = None

@app.get("/api/settings")
def get_system_settings(db: Session = Depends(get_db)):
    url_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_url"))
    key_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_key"))
    
    raw_key = key_setting.value if key_setting else ""
    masked_key = ""
    if raw_key:
        if len(raw_key) > 10:
            masked_key = raw_key[:6] + "..." + raw_key[-4:]
        else:
            masked_key = "********"
            
    return {
        "success": True,
        "data": {
            "supabase_url": url_setting.value if url_setting else "",
            "supabase_key": masked_key,
            "has_key": bool(raw_key)
        }
    }

@app.post("/api/settings")
def update_system_settings(cfg: SettingsUpdate, db: Session = Depends(get_db)):
    if cfg.supabase_url is not None:
        url_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_url"))
        if not url_setting:
            url_setting = SystemSetting(key="supabase_url")
            db.add(url_setting)
        url_setting.value = cfg.supabase_url.strip()
        
    if cfg.supabase_key is not None:
        raw_key = cfg.supabase_key.strip()
        if raw_key and not raw_key.endswith("..."):
            key_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_key"))
            if not key_setting:
                key_setting = SystemSetting(key="supabase_key")
                db.add(key_setting)
            key_setting.value = raw_key
        elif not raw_key:
            key_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_key"))
            if key_setting:
                key_setting.value = ""
                
    db.commit()
    return {"success": True, "message": "Cấu hình hệ thống đã được cập nhật thành công!"}

@app.post("/api/admin/login")
def admin_login(p: AdminLogin, db: Session = Depends(get_db)):
    import hashlib
    password_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "admin_password"))
    default_hash = hashlib.sha256("admin".encode()).hexdigest()
    stored_hash = password_setting.value if password_setting else default_hash
    
    input_hash = hashlib.sha256(p.password.encode()).hexdigest()
    if input_hash == stored_hash:
        return {"success": True, "token": "xiaozhi-admin-session-token"}
    else:
        return {"success": False, "message": "Mật khẩu quản trị viên không chính xác!"}

@app.post("/api/admin/change-password")
def admin_change_password(p: AdminChangePassword, db: Session = Depends(get_db)):
    import hashlib
    password_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "admin_password"))
    default_hash = hashlib.sha256("admin".encode()).hexdigest()
    stored_hash = password_setting.value if password_setting else default_hash
    
    current_hash = hashlib.sha256(p.current_password.encode()).hexdigest()
    if current_hash != stored_hash:
        return {"success": False, "message": "Mật khẩu hiện tại không chính xác!"}
    
    new_hash = hashlib.sha256(p.new_password.encode()).hexdigest()
    if not password_setting:
        password_setting = SystemSetting(key="admin_password")
        db.add(password_setting)
    password_setting.value = new_hash
    db.commit()
    return {"success": True, "message": "Đã đổi mật khẩu quản trị viên thành công!"}

async def sync_device_to_supabase_bg(sb_url: str, sb_key: str, payload: dict):
    import httpx
    import logging
    logger = logging.getLogger(__name__)
    
    endpoint = f"{sb_url}/rest/v1/devices"
    headers = {
        "apikey": sb_key,
        "Authorization": f"Bearer {sb_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates"
    }
    
    async with httpx.AsyncClient(timeout=10.0) as client:
        try:
            res = await client.post(endpoint, json=payload, headers=headers)
            if res.status_code in (200, 201):
                logger.info("Successfully synced device to Supabase in background.")
            else:
                logger.error(f"Supabase background sync failed with status {res.status_code}: {res.text}")
        except Exception as e:
            logger.error(f"Error in Supabase background sync: {e}")

def queue_supabase_sync(device: Device, db: Session, background_tasks: BackgroundTasks):
    import logging
    logger = logging.getLogger(__name__)
    
    url_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_url"))
    key_setting = db.scalar(select(SystemSetting).where(SystemSetting.key == "supabase_key"))
    
    sb_url = url_setting.value.strip() if url_setting and url_setting.value else ""
    sb_key = key_setting.value.strip() if key_setting and key_setting.value else ""
    
    # Fallback to env
    if not sb_url:
        from api.app.core.config import get_settings
        config_settings = get_settings()
        sb_url = config_settings.supabase_url.strip() if config_settings.supabase_url else ""
        sb_key = config_settings.supabase_key.strip() if config_settings.supabase_key else ""
        
    if not sb_url or not sb_key:
        logger.info("Supabase sync skipped: URL or Key is not configured.")
        return
        
    payload = {
        "external_id": device.external_id,
        "name": device.name,
        "status": device.status,
        "mac_address": device.mac_address,
        "current_language": device.current_language,
        "current_voice": device.current_voice,
        "ai_prompt_template": device.ai_prompt_template,
        "asr_speed": device.asr_speed,
        "tts_speech_speed": device.tts_speech_speed,
        "tts_pitch": device.tts_pitch,
        "board_name": device.board_name,
        "serial_number": device.serial_number,
        "is_auth": device.is_auth,
        "app_version": device.app_version,
        "last_seen_at": device.last_seen_at.isoformat() if device.last_seen_at else None
    }
    
    background_tasks.add_task(sync_device_to_supabase_bg, sb_url, sb_key, payload)

@app.post("/api/devices/activate")
async def activate_device(act: DeviceActivate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    import random
    import httpx
    import logging
    logger = logging.getLogger(__name__)

    code = act.code.strip()
    if not code.isdigit() or len(code) != 6:
        raise HTTPException(status_code=400, detail="Mã kích hoạt phải là dãy 6 chữ số")
    
    device_name = act.name.strip() if act.name and act.name.strip() else f"Loa Globy KH-{code[-4:]}"

    # Query the default preset from DB
    default_preset = db.scalars(select(Preset)).first()
    if default_preset:
        llm_model = default_preset.llm_model or "qwen"
        lang = default_preset.language or "vi"
        voice = default_preset.tts_voice or "vi-VN-HoaiMyNeural"
        character = default_preset.character_prompt or "Bạn là trợ lý đắc lực của bé..."
        asr_speed = default_preset.asr_speed or "normal"
        tts_speech_speed = default_preset.tts_speech_speed or "normal"
        tts_pitch = default_preset.tts_pitch or 0
        mcp_endpoints = json.loads(default_preset.mcp_endpoints_json) if default_preset.mcp_endpoints_json else ["2", "104"]
    else:
        llm_model = "qwen"
        lang = "vi"
        voice = "vi-VN-HoaiMyNeural"
        character = "Bạn là trợ lý đắc lực của bé..."
        asr_speed = "normal"
        tts_speech_speed = "normal"
        tts_pitch = 0
        mcp_endpoints = ["2", "104"]

    # 1. Require an active account
    active_accounts = db.scalars(
        select(XiaozhiAccount).where(
            XiaozhiAccount.is_active == True
        )
    ).all()

    if not active_accounts:
        raise HTTPException(
            status_code=400,
            detail="Chưa có tài khoản Xiaozhi nào được thêm hoặc kích hoạt. Vui lòng thêm tài khoản quản trị trước."
        )

    account = active_accounts[0]

    # 2. Check if it's a mock token for testing/demo purposes
    if account.bearer_token == "mock" or account.bearer_token.startswith("mock_"):
        mac_suffix = f"{random.randint(16, 255):02X}"
        new_mac = f"AA:BB:CC:DD:EE:{mac_suffix}"
        new_id = f"agent-act-{random.randint(1000, 9999)}"
        
        existing = db.scalar(select(Device).where(Device.mac_address == new_mac))
        if existing:
            new_mac = f"AA:BB:CC:DD:EE:{random.randint(16, 255):02X}"
        
        db_device = Device(
            external_id=new_id,
            name=device_name,
            status="online",
            mac_address=new_mac,
            current_language=lang,
            current_voice=voice,
            ai_prompt_template=character,
            asr_speed=asr_speed,
            tts_speech_speed=tts_speech_speed,
            tts_pitch=tts_pitch,
            mcp_endpoints_json=json.dumps(mcp_endpoints),
            board_name="jiuchuan-s3",
            serial_number=f"MOCK_SERIAL_{random.randint(100000, 999999)}",
            is_auth=True,
            app_version="2.1.6",
            last_seen_at=func.now()
        )
        db.add(db_device)
        db.commit()
        db.refresh(db_device)
        await sync_device_to_supabase(db_device, db)
        
        return {
            "success": True, 
            "mac_address": new_mac, 
            "name": db_device.name,
            "message": f"Kích hoạt thiết bị {db_device.name} (Giả lập) thành công!"
        }

    # 3. Call the real Xiaozhi API for activation
    headers = {
        "Authorization": f"Bearer {account.bearer_token.strip()}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0"
    }
    
    async with httpx.AsyncClient(timeout=15.0) as client:
        try:
            res = await client.post(
                "https://xiaozhi.me/api/agents/devices",
                json={"verificationCode": code},
                headers=headers
            )
            
            if res.status_code == 200:
                res_data = res.json()
                if res_data.get("success") or res_data.get("code") == "ADD_DEVICE_SUCCESS":
                    device_info = res_data.get("data", {})
                    real_mac = device_info.get("mac_address") or device_info.get("macAddress")
                    real_id = str(device_info.get("id"))
                    agent_id = str(device_info.get("agent_id") or "")
                    board_name = device_info.get("board_name")
                    serial_number = device_info.get("serial_number")
                    is_auth = device_info.get("is_auth", False)
                    app_version = device_info.get("app_version")

                    if not real_mac or not real_id:
                        raise HTTPException(
                            status_code=400,
                            detail="Kích hoạt thành công nhưng phản hồi từ Xiaozhi thiếu thông tin thiết bị (MAC address hoặc ID)."
                        )

                    # Save device to DB
                    db_device = db.scalar(select(Device).where(Device.external_id == real_id))
                    if not db_device:
                        db_device = Device(external_id=real_id)
                        db.add(db_device)
                    
                    db_device.name = f"Globy-{real_id}"
                    db_device.mac_address = real_mac
                    db_device.status = "online"
                    db_device.account_id = account.id
                    db_device.current_language = lang
                    db_device.current_voice = voice
                    db_device.ai_prompt_template = character
                    db_device.asr_speed = asr_speed
                    db_device.tts_speech_speed = tts_speech_speed
                    db_device.tts_pitch = tts_pitch
                    db_device.mcp_endpoints_json = json.dumps(mcp_endpoints)
                    
                    # Hardware metadata fields
                    db_device.board_name = board_name
                    db_device.serial_number = serial_number
                    db_device.is_auth = is_auth
                    db_device.app_version = app_version
                    db_device.last_seen_at = func.now()
                    
                    db.commit()
                    db.refresh(db_device)
                    queue_supabase_sync(db_device, db, background_tasks)

                    # Sync preset configuration to Xiaozhi
                    if agent_id:
                        config_url = f"https://xiaozhi.me/api/agents/{agent_id}/config"
                        
                        model_mapping = {
                            "xz-lite": "xz-lite",
                            "qwen": "qwen",
                            "deepseek": "deepseek",
                            "doubao": "doubao",
                            "gpt-5": "gpt-5"
                        }
                        mapped_model = model_mapping.get(llm_model, "qwen")

                        config_payload = {
                            "agent_name": f"Globy-{real_id}",
                            "llm_model": mapped_model,
                            "tts_voice": voice,
                            "tts_speech_speed": tts_speech_speed,
                            "tts_pitch": tts_pitch,
                            "asr_speed": asr_speed,
                            "language": lang,
                            "character": character,
                            "memory_type": "LONG_TERM",
                            "mcp_endpoints": mcp_endpoints,
                            "knowledge_base_ids": []
                        }
                        try:
                            await client.post(config_url, json=config_payload, headers=headers)
                        except Exception as config_err:
                            logger.error(f"Lỗi đẩy cấu hình mặc định lên Xiaozhi: {config_err}")

                    return {
                        "success": True, 
                        "mac_address": real_mac, 
                        "name": f"Globy-{real_id}",
                        "message": f"Kích hoạt thiết bị Globy-{real_id} trên Xiaozhi thành công!"
                    }
                else:
                    msg = res_data.get("message") or ""
                    if "验证码" in msg or res_data.get("code") == "INVALID_VERIFICATION_CODE":
                        detail_msg = "Mã kích hoạt không hợp lệ hoặc đã hết hạn. Vui lòng kiểm tra lại mã trên loa."
                    else:
                        detail_msg = f"Kích hoạt thất bại: {msg}" if msg else "Kích hoạt thất bại từ máy chủ Xiaozhi."
                    raise HTTPException(status_code=400, detail=detail_msg)
            else:
                try:
                    res_data = res.json()
                    msg = res_data.get("message") or ""
                    if "验证码" in msg or res_data.get("code") == "INVALID_VERIFICATION_CODE":
                        detail_msg = "Mã kích hoạt không hợp lệ hoặc đã hết hạn. Vui lòng kiểm tra lại mã trên loa."
                    else:
                        detail_msg = f"Lỗi từ máy chủ Xiaozhi: {msg}" if msg else f"Lỗi máy chủ Xiaozhi (Status: {res.status_code})"
                except Exception:
                    detail_msg = f"Lỗi máy chủ Xiaozhi (Status: {res.status_code})"
                
                raise HTTPException(status_code=400, detail=detail_msg)
                
        except HTTPException:
            raise
        except Exception as real_err:
            logger.error(f"Lỗi kích hoạt thật: {real_err}")
            raise HTTPException(
                status_code=502,
                detail=f"Không thể kết nối đến máy chủ Xiaozhi để kích hoạt: {str(real_err)}"
            )


